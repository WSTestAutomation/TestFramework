# coding=utf-8

import os
import csv
import openpyxl
from xlrd import open_workbook


def rewrite_csv_name(file_path, insurance_no):
    file = os.listdir(file_path)
    old_name = os.path.join(file_path, file[0])
    new_data_path = os.path.join(file_path, '{}-00.xlsx'.format(insurance_no))
    os.rename(old_name, new_data_path)
    return new_data_path


def read_csv_file(file_path, sheet=0, title_line=True):
    _data = list()  # 用于存储每行生成的数据。
    if not _data:
        workbook = open_workbook(file_path)
        if type(sheet) not in [int, str]:
            raise Exception('Please pass in <type int> or <type str>, not {0}'.format(type(sheet)))
        elif type(sheet) == int:
            s = workbook.sheet_by_index(sheet)
        else:
            s = workbook.sheet_by_name(sheet)

        if title_line:
            title = s.row_values(0)  # 首行为title
            for col in range(1, s.nrows):
                # 依次遍历其余行，与首行组成dict，拼到self._data中
                _data.append(dict(zip(title, s.row_values(col))))
        else:
            for col in range(0, s.nrows):
                # 遍历所有行，拼到self._data中
                _data.append(s.row_values(col))
    return _data


def modify_csv_file(file_path, modify_cell, modify_value, sheet=0):
    work_book = openpyxl.load_workbook(file_path)
    work_sheet = work_book.worksheets[sheet]
    work_sheet[modify_cell] = modify_value
    work_book.save(file_path)


def create_csv_file(file_path, csv_header, csv_data, sheet_name='sheet1'):
    work_book = openpyxl.Workbook()
    work_sheet = work_book.active
    work_sheet.title = sheet_name

    for i in range(len(csv_header)):
        work_sheet.cell(1, i + 1, csv_header[i])

    work_sheet.append(csv_data)
    work_book.save(file_path)


def get_cell_value_by_position(file_path, cell_position, sheet=0):
    work_book = openpyxl.load_workbook(file_path)
    work_sheet = work_book.worksheets[sheet]
    print(work_sheet[cell_position].value)
    return work_sheet[cell_position].value


def write_csv_person_info(info, csv_file):
    with open(csv_file, 'w', encoding='utf-8', newline='') as f:
        fieldnames = ['name', 'id', 'phone', 'account', 'password']
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerow(info)


def write_csv_file(file_path, cell, value, sheet=0):
    work_book = openpyxl.load_workbook(file_path)
    work_sheet = work_book.worksheets[sheet]
    work_sheet[cell] = value
    work_book.save(file_path)


def read_csv_person_info(csv_file):
    with open(csv_file, encoding='utf-8', newline='') as f:
        reader = csv.DictReader(f)
        return [row for row in reader]


def rewrite_csv_person_info(account, csv_file):
    info = read_csv_person_info(csv_file)[0]
    info['account'] = account
    with open(csv_file, 'w', encoding='utf-8', newline='') as f:
        fieldnames = ['name', 'id', 'phone', 'account', 'password']
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerow(info)

"""
if __name__ == '__main__':
    base_dir = os.path.dirname(os.path.dirname(__file__))
    test_data_path = os.path.join(base_dir, 'test_data')
    #nput_file = os.path.join(test_data_path, 'input_data.xlsx')
    rewrite_csv_name(input_file, 'asd')
    read_csv_file(input_file)
""" 